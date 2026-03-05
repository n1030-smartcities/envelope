"""
db_builder.py — Constrói (ou reconstrói) o banco SQLite de parâmetros
a partir dos XLSX de cada cidade.

Uso:
    python db_builder.py                           # gera params.db na pasta do script
    python db_builder.py --data data/ --out params.db
    python db_builder.py --force                   # reconstrói mesmo se já existe

Schema SQLite:
    zonas(cidade, chave, sigla_display, nome,
          ca_basico, ca_permissivel, gabarito_max_m,
          taxa_ocupacao, taxa_permeabilidade, recuo_frontal_m,
          af_divisor, af_acrescimo, af_minimo, af_facultado_ate_m,
          emb_permitido, emb_altura_max_m, emb_to, emb_pode_lateral,
          emb_pode_fundos, emb_pct_perimetro,
          fracao_minima_m2, observacoes)
"""

from __future__ import annotations
import argparse
import glob
import os
import re
import sqlite3
from typing import Optional
import openpyxl

# ─────────────────────────────────────────────────────────────
# DDL
# ─────────────────────────────────────────────────────────────
CREATE_TABLE = """
CREATE TABLE IF NOT EXISTS zonas (
    cidade              TEXT NOT NULL,
    chave               TEXT NOT NULL,
    sigla_display       TEXT NOT NULL,
    nome                TEXT NOT NULL,
    ca_basico           REAL NOT NULL,
    ca_permissivel      REAL,
    gabarito_max_m      REAL,
    taxa_ocupacao       REAL NOT NULL DEFAULT 0.5,
    taxa_permeabilidade REAL NOT NULL DEFAULT 0.2,
    recuo_frontal_m     REAL NOT NULL DEFAULT 5.0,
    af_divisor          REAL,
    af_acrescimo        REAL NOT NULL DEFAULT 0.0,
    af_minimo           REAL NOT NULL DEFAULT 0.0,
    af_facultado_ate_m  REAL NOT NULL DEFAULT 0.0,
    emb_permitido       INTEGER NOT NULL DEFAULT 0,
    emb_altura_max_m    REAL NOT NULL DEFAULT 9.0,
    emb_to              REAL,
    emb_pode_lateral    INTEGER NOT NULL DEFAULT 0,
    emb_pode_fundos     INTEGER NOT NULL DEFAULT 0,
    emb_pct_perimetro   REAL NOT NULL DEFAULT 0.0,
    fracao_minima_m2    REAL,
    observacoes         TEXT NOT NULL DEFAULT '',
    PRIMARY KEY (cidade, chave)
);
"""

_ALTURA_PAV_M = 3.0   # conversão pavimentos→metros (Curitiba)


# ─────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────

def _sf(v) -> Optional[float]:
    if v is None: return None
    try: return float(str(v).replace(",", ".").strip())
    except: return None

def _parse_to(v) -> Optional[float]:
    if v is None: return None
    s = str(v).strip()
    m = re.search(r"^(\d+(?:[.,]\d+)?)", s)
    return float(m.group(1).replace(",", ".")) / 100.0 if m else None

def _parse_recuo(v) -> Optional[float]:
    if v is None: return None
    s = str(v).strip().lower()
    if "alinhamento" in s: return 0.0
    m = re.search(r"(\d+(?:[.,]\d+)?)", s)
    return float(m.group(1).replace(",", ".")) if m else None

def _parse_afas_texto(texto: str, gabarito_pav: Optional[int]):
    """
    Interpreta texto da coluna 'Afastamento Divisa' de Curitiba.
    Retorna (divisor, acrescimo, minimo, facultado_ate_m).
    """
    if not texto or str(texto).strip() in ("", "None"):
        return None, 0.0, 0.0, 0.0
    s = str(texto).strip()
    try:
        val = float(s.replace(",", "."))
        return None, 0.0, val, 0.0   # fixo
    except ValueError:
        pass
    m_fac = re.search(r"até\s+(\d+)\s+pav", s, re.IGNORECASE)
    m_div = re.search(r"H/(\d+)", s, re.IGNORECASE)
    m_min = re.search(r"mín\s*([\d,\.]+)", s, re.IGNORECASE)
    divisor   = float(m_div.group(1)) if m_div else 6.0
    minimo    = float(m_min.group(1).replace(",", ".")) if m_min else 2.5
    fac_m     = int(m_fac.group(1)) * _ALTURA_PAV_M if m_fac else 0.0
    return divisor, 0.0, minimo, fac_m

def _parse_fracao(obs: str) -> Optional[float]:
    if not obs: return None
    m = re.search(r"[Ff]ra[çc][aã]o\s+m[íi]n\s*([\d]+)", obs)
    return float(m.group(1)) if m else None


# ─────────────────────────────────────────────────────────────
# Loader Curitiba
# ─────────────────────────────────────────────────────────────

def _load_curitiba(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Zoneamento Curitiba"]
    raw: dict[str, dict] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        sigla = row[1]
        if not sigla: continue
        sigla = str(sigla).strip()
        nome  = str(row[0] or sigla).strip()

        ca_b  = _sf(row[6])
        ca_p  = _sf(row[7])
        alt_b = _sf(row[8])    # pavimentos
        to_v  = _parse_to(row[12])
        recuo = _parse_recuo(row[13])
        perm  = _sf(row[14])
        afas_t = str(row[15] or "").strip()
        obs   = str(row[17] or "").strip()

        gab_m: Optional[float] = None
        gab_pav: Optional[int] = None
        if alt_b is not None:
            try:
                gab_pav = int(alt_b)
                gab_m = alt_b * _ALTURA_PAV_M
            except: pass

        if sigla not in raw:
            raw[sigla] = dict(nome=nome, ca_basico=None, ca_permissivel=None,
                              gabarito_max_m=None, gabarito_pav=None,
                              taxa_ocupacao=None, recuo_frontal_m=None,
                              taxa_permeabilidade=None, afas_texto="",
                              fracao_minima_m2=None, observacoes="")
        acc = raw[sigla]
        acc["nome"] = nome
        if ca_b and (acc["ca_basico"] is None or ca_b > acc["ca_basico"]):
            acc["ca_basico"] = ca_b
        if ca_p and (acc["ca_permissivel"] is None or ca_p > (acc["ca_permissivel"] or 0)):
            acc["ca_permissivel"] = ca_p
        if gab_m and (acc["gabarito_max_m"] is None or gab_m > acc["gabarito_max_m"]):
            acc["gabarito_max_m"] = gab_m
            acc["gabarito_pav"] = gab_pav
        if to_v  and acc["taxa_ocupacao"]       is None: acc["taxa_ocupacao"] = to_v
        if recuo is not None and acc["recuo_frontal_m"] is None: acc["recuo_frontal_m"] = recuo
        if perm  and acc["taxa_permeabilidade"] is None: acc["taxa_permeabilidade"] = perm / 100.0
        if afas_t and not acc["afas_texto"]: acc["afas_texto"] = afas_t
        if obs and not acc["observacoes"]:
            acc["observacoes"] = obs
            frac = _parse_fracao(obs)
            if frac and acc["fracao_minima_m2"] is None: acc["fracao_minima_m2"] = frac

    rows = []
    for sigla, acc in raw.items():
        if acc["ca_basico"] is None: continue
        div, acr, mini, fac = _parse_afas_texto(acc["afas_texto"], acc.get("gabarito_pav"))
        rows.append(dict(
            cidade="curitiba", chave=sigla, sigla_display=sigla, nome=acc["nome"],
            ca_basico=acc["ca_basico"], ca_permissivel=acc["ca_permissivel"],
            gabarito_max_m=acc["gabarito_max_m"],
            taxa_ocupacao=acc["taxa_ocupacao"] or 0.50,
            taxa_permeabilidade=acc["taxa_permeabilidade"] or 0.25,
            recuo_frontal_m=acc["recuo_frontal_m"] or 5.0,
            af_divisor=div, af_acrescimo=acr, af_minimo=mini, af_facultado_ate_m=fac,
            emb_permitido=0, emb_altura_max_m=9.0, emb_to=None,
            emb_pode_lateral=0, emb_pode_fundos=0, emb_pct_perimetro=0.0,
            fracao_minima_m2=acc["fracao_minima_m2"], observacoes=acc["observacoes"],
        ))
    return rows


# ─────────────────────────────────────────────────────────────
# Loader Joinville
# ─────────────────────────────────────────────────────────────

def _ler_aba2(ws, val_col: int, extra: dict = None):
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        mz, st = row[0], row[1]
        if mz is None or st is None: continue
        entry = {"value": _sf(row[val_col])}
        if extra:
            for name, idx in extra.items():
                entry[name] = row[idx]
        result[(str(mz).strip(), str(st).strip())] = entry
    return result

def _load_joinville(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    cal  = _ler_aba2(wb["CAL"],      2)
    gab  = _ler_aba2(wb["GABARITO"], 2)
    to_  = _ler_aba2(wb["TO"],       2)
    rec  = _ler_aba2(wb["RECUO"],    2)
    rlf  = _ler_aba2(wb["REC.L.F"],  2, {"divisor":3,"minimo":4,"acrescimo":5})
    tp_  = _ler_aba2(wb["TP"],       2)
    qa_  = _ler_aba2(wb["Q.A"],      2)
    emb  = _ler_aba2(wb["EMBASAMENTO"], 2)

    mz_nomes = {}
    for row in wb["MACROZONAS"].iter_rows(min_row=2, values_only=True):
        if row[1]: mz_nomes[str(row[1]).strip()] = str(row[4] or row[1]).strip()
    st_nomes = {}
    for row in wb["SETORES"].iter_rows(min_row=2, values_only=True):
        if row[1]: st_nomes[str(row[1]).strip()] = str(row[2] or row[1]).strip()

    rows = []
    for (mz, st) in cal:
        ca_e = cal[(mz, st)]
        if ca_e["value"] is None: continue

        gab_e = gab.get((mz, st))
        to_e  = to_.get((mz, st))
        rec_e = rec.get((mz, st))
        tp_e  = tp_.get((mz, st)) or tp_.get((mz, "FR"))
        qa_e  = qa_.get((mz, st))
        rlf_e = rlf.get((mz, st))
        emb_e = emb.get((mz, st))

        div = acr = mini = None
        if rlf_e:
            div  = _sf(rlf_e.get("divisor"))
            mini = _sf(rlf_e.get("minimo")) or 0.0
            acr  = _sf(rlf_e.get("acrescimo")) or 0.0

        emb_perm = 0; emb_to = None
        if emb_e and emb_e["value"]:
            emb_perm = 1
            emb_to   = emb_e["value"] / 100.0

        chave = f"{mz}|{st}"
        nome  = f"{mz_nomes.get(mz,mz)} — {st_nomes.get(st,st)}"
        rows.append(dict(
            cidade="joinville", chave=chave, sigla_display=f"{mz} / {st}", nome=nome,
            ca_basico=float(ca_e["value"]), ca_permissivel=None,
            gabarito_max_m=gab_e["value"] if gab_e else None,
            taxa_ocupacao=(to_e["value"]/100.0) if to_e and to_e["value"] else 0.60,
            taxa_permeabilidade=(tp_e["value"]/100.0) if tp_e and tp_e["value"] else 0.20,
            recuo_frontal_m=float(rec_e["value"]) if rec_e and rec_e["value"] else 5.0,
            af_divisor=div, af_acrescimo=acr or 0.0, af_minimo=mini or 0.0,
            af_facultado_ate_m=0.0,
            emb_permitido=emb_perm, emb_altura_max_m=9.0, emb_to=emb_to,
            emb_pode_lateral=1 if emb_perm else 0,
            emb_pode_fundos=1 if emb_perm else 0,
            emb_pct_perimetro=0.50 if emb_perm else 0.0,
            fracao_minima_m2=qa_e["value"] if qa_e else None,
            observacoes="",
        ))
    return rows


# ─────────────────────────────────────────────────────────────
# Builder
# ─────────────────────────────────────────────────────────────

def build_db(data_dir: str, db_path: str, force: bool = False):
    if os.path.exists(db_path) and not force:
        print(f"[db_builder] {db_path} já existe. Use --force para reconstruir.")
        return

    con = sqlite3.connect(db_path)
    con.execute("DROP TABLE IF EXISTS zonas")
    con.execute(CREATE_TABLE)

    loaders = {
        "curitiba": ("parametros_curitiba_2026.xlsx", _load_curitiba),
        "joinville": ("parametros_joinville_2026.xlsx", _load_joinville),
    }
    # Aceita qualquer arquivo parametros_{cidade}*.xlsx
    for cidade, (default_name, loader) in loaders.items():
        candidates = sorted(glob.glob(os.path.join(data_dir, f"parametros_{cidade}*.xlsx")))
        path = candidates[-1] if candidates else os.path.join(data_dir, default_name)
        if not os.path.exists(path):
            print(f"[db_builder] XLSX não encontrado: {path} — pulando {cidade}")
            continue
        rows = loader(path)
        cols = [c for c in CREATE_TABLE.split("\n")
                if c.strip() and not c.strip().startswith(("CREATE","PRIMARY",")"))]
        col_names = [r.strip().split()[0] for r in rows[0].keys()] if rows else []
        con.executemany(
            f"INSERT OR REPLACE INTO zonas VALUES ({','.join('?' for _ in rows[0])})",
            [tuple(r[k] for k in r) for r in rows]
        )
        print(f"[db_builder] {cidade}: {len(rows)} zonas inseridas de {os.path.basename(path)}")

    con.commit()
    con.close()
    print(f"[db_builder] Banco salvo em {db_path}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--data",  default="data", help="Pasta com os XLSX")
    ap.add_argument("--out",   default="params.db", help="Caminho do SQLite")
    ap.add_argument("--force", action="store_true")
    args = ap.parse_args()
    build_db(args.data, args.out, args.force)
